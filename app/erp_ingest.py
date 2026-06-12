from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from app.erp import INPUT_HEADER_ALIASES

EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff", ".heic"}

CANONICAL_HEADERS = (
    "产品名称",
    "颜色",
    "尺寸",
    "价格",
    "重量",
    "供应商",
)

SIZE_PATTERN = re.compile(r"^\d+[*+×xX]\d+(?:/\d+[*+×xX]\d+)?(?:套)?$", re.I)
SIZE_PRICE_ROW = re.compile(
    r"^(\d+[*+×xX]\d+(?:/\d+[*+×xX]\d+)?(?:套)?)\s+(\d+(?:\.\d+)?)(?:\s+(\d+(?:\.\d+)?))?",
    re.I,
)
SIZE_DIM_FIX = re.compile(r"(\d)[+×xX](\d)")
SKIP_LINE_KEYWORDS = ("定制", "按照")

PRICE_SHEET_HEADER_ALIASES = {
    "size": ("产品尺寸", "尺寸"),
    "price": ("成本价", "价格", "单价"),
    "weight": ("单品重量", "单品重量kg", "单品重量(kg)", "重量(kg)", "重量kg", "重量"),
}

NOISE_LINE_PATTERNS = (
    re.compile(r"网盘"),
    re.compile(r"pan\.baidu", re.I),
    re.compile(r"https?://", re.I),
    re.compile(r"链接\s*[:：]"),
    re.compile(r"提取码"),
    re.compile(r"pwd\s*=", re.I),
    re.compile(r"来自百度"),
    re.compile(r"分享的文件"),
    re.compile(r"超级会员"),
    re.compile(r"通过网盘"),
    re.compile(r"镜界摄影"),
    re.compile(r"境界摄影"),
)


def ingest_upload(filename: str, content: bytes) -> tuple[str, str, str]:
    name = (filename or "").strip()
    suffix = _suffix(name)
    if suffix in EXCEL_EXTENSIONS:
        return parse_excel_bytes(content), "excel_upload", name
    if suffix in IMAGE_EXTENSIONS:
        return parse_image_ocr(content), "image_ocr", name
    raise ValueError("仅支持 Excel（.xlsx）或截图（.png/.jpg/.webp 等）")


def parse_excel_bytes(content: bytes) -> str:
    try:
        import openpyxl
    except ImportError as exc:
        raise ValueError("缺少 openpyxl 依赖，无法解析 Excel") from exc

    workbook = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        blocks: list[str] = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = [
                [_cell_text(cell) for cell in row]
                for row in sheet.iter_rows(values_only=True)
                if any(_cell_text(cell) for cell in row)
            ]
            if not rows:
                continue
            text = _plain_rows_to_text(rows) or _rows_to_table_text(rows)
            if text:
                blocks.append(text)
        if not blocks:
            raise ValueError("Excel 里没有可识别的表格数据")
        return "\n".join(blocks)
    finally:
        workbook.close()


def parse_image_ocr(content: bytes) -> str:
    try:
        from PIL import Image
    except ImportError as exc:
        raise ValueError("缺少 Pillow 依赖，无法识别截图") from exc

    image = Image.open(BytesIO(content))
    tokens = _ocr_tokens(image)
    if not tokens:
        raise ValueError("截图未识别到文字，请换清晰图片或改粘贴表格文本")
    text = sanitize_ocr_text(_tokens_to_table_text(tokens))
    if not text.strip():
        raise ValueError("截图 OCR 结果为空，请换更清晰图片，或在下方粘贴识别文本")
    return text


def sanitize_ocr_text(text: str) -> str:
    lines = [line.rstrip() for line in str(text or "").splitlines() if line.strip()]
    kept = [line for line in lines if not _is_noise_line(line)]
    return "\n".join(kept)


def normalize_raw_text_for_erp(text: str) -> tuple[str, dict[str, str]]:
    cleaned = _repair_ocr_lines(sanitize_ocr_text(str(text or "").strip()))
    if not cleaned:
        return "", {}

    sheet = parse_supplier_price_sheet(cleaned)
    if sheet:
        return sheet

    return cleaned, extract_price_sheet_hints(cleaned)


def parse_supplier_price_sheet(text: str) -> tuple[str, dict[str, str]] | None:
    lines = [line.rstrip() for line in sanitize_ocr_text(text).splitlines() if line.strip()]
    if not lines:
        return None

    hints = extract_price_sheet_hints("\n".join(lines))
    colors = _extract_colors(text)
    supplier = hints.get("supplier", "")
    product_name = hints.get("product_name", "")
    size_rows = _extract_size_price_rows(lines)

    if len(size_rows) < 1 or len(colors) < 1:
        return None

    canonical_product_name = product_name or hints.get("product_name", "")
    output = ["\t".join(CANONICAL_HEADERS)]
    for size, price, weight in size_rows:
        for color in colors:
            output.append(
                "\t".join(
                    [
                        canonical_product_name,
                        color,
                        size,
                        price,
                        weight,
                        supplier,
                    ]
                )
            )
    return "\n".join(output), hints


def _extract_colors(text: str) -> list[str]:
    for line in text.splitlines():
        if _is_noise_line(line):
            continue
        color_match = re.search(r"颜色[：:]\s*(.+)", line.strip())
        if color_match:
            colors = [part.strip() for part in re.split(r"[、,，/]", color_match.group(1)) if part.strip()]
            return [color for color in colors if color and not _is_bad_product_name(color)]
    return []


def extract_price_sheet_hints(text: str) -> dict[str, str]:
    hints: dict[str, str] = {}
    lines = [line.strip() for line in text.splitlines() if line.strip() and not _is_noise_line(line)]

    product_name = _pick_product_name(lines)
    if product_name:
        hints["product_name"] = product_name

    for line in lines:
        supplier_match = re.search(r"供应商[：:]\s*(.+)", line)
        if supplier_match:
            hints["supplier"] = supplier_match.group(1).strip()

        width_match = re.search(r"(?:门幅[^0-9]*|度)(\d+(?:\.\d+)?)\s*米", line)
        if width_match:
            hints["width_span"] = f"{width_match.group(1)}米"

    return hints


def _pick_product_name(lines: list[str]) -> str:
    candidates: list[tuple[int, str]] = []
    for line in lines:
        if "沙发垫" not in line:
            continue
        name = _extract_product_name(line)
        if not name or _is_bad_product_name(name):
            continue
        score = 0
        if "底价" in line or "报价" in line or "价格表" in line:
            score += 20
        if re.search(r"^[^：:\s]{1,8}[：:]", line):
            score += 10
        if len(name) <= 12:
            score += 5
        if re.search(r"\(.*?\)", name) or re.search(r"（.*?）", name):
            score += 3
        candidates.append((score, name))
    if not candidates:
        return ""
    candidates.sort(key=lambda item: (-item[0], len(item[1])))
    return candidates[0][1]


def _is_noise_line(line: str) -> bool:
    cleaned = line.strip()
    if not cleaned:
        return True
    return any(pattern.search(cleaned) for pattern in NOISE_LINE_PATTERNS)


def _is_bad_product_name(name: str) -> bool:
    if re.search(r"摄影|网盘|分享|http|pan\.|pwd=", name, re.I):
        return True
    if re.search(r"\d{4}[-/年]\d", name):
        return True
    if re.search(r"\d+色沙发垫", name):
        return True
    return False


def _extract_product_name(line: str) -> str:
    match = re.search(r"(.+?沙发垫)", line)
    if not match:
        return ""
    name = match.group(1).strip()
    name = re.sub(r"(底价|报价|价格表)$", "", name).strip()
    name = re.sub(r"^[^：:]*[：:]\s*", "", name).strip()
    return name


def _normalize_size_text(size: str) -> str:
    cleaned = str(size or "").strip()
    if not cleaned:
        return ""
    while SIZE_DIM_FIX.search(cleaned):
        cleaned = SIZE_DIM_FIX.sub(r"\1*\2", cleaned)
    return cleaned


def _matches_size(text: str) -> bool:
    return bool(SIZE_PATTERN.match(_normalize_size_text(text)))


def _ocr_text_has_price_data(text: str) -> bool:
    for line in text.splitlines():
        if _is_noise_line(line):
            continue
        if any(keyword in line for keyword in SKIP_LINE_KEYWORDS):
            continue
        parsed = _parse_size_price_cells(_split_table_cells(line))
        if parsed and str(parsed[1] or "").strip():
            return True
    return False


def _repair_ocr_lines(text: str) -> str:
    lines = [line.rstrip() for line in str(text or "").splitlines() if line.strip()]
    if not lines:
        return ""

    repaired: list[str] = []
    index = 0
    while index < len(lines):
        line = lines[index]
        if _is_noise_line(line) or any(keyword in line for keyword in SKIP_LINE_KEYWORDS):
            repaired.append(line)
            index += 1
            continue

        parsed = _parse_size_price_cells(_split_table_cells(line))
        if parsed and parsed[0] and not (str(parsed[1] or "").strip() or str(parsed[2] or "").strip()):
            numbers: list[str] = []
            cursor = index + 1
            while cursor < len(lines) and len(numbers) < 2:
                next_cells = _split_table_cells(lines[cursor])
                if not next_cells:
                    cursor += 1
                    continue
                if all(re.fullmatch(r"\d+(?:\.\d+)?", cell) for cell in next_cells):
                    numbers.extend(next_cells[: 2 - len(numbers)])
                    cursor += 1
                    continue
                break
            if numbers:
                price, weight = _normalize_price_weight(numbers[0], numbers[1] if len(numbers) > 1 else "")
                repaired.append("\t".join([parsed[0], price, weight]))
                index = cursor
                continue

        repaired.append(line)
        index += 1
    return "\n".join(repaired)


def _extract_size_price_rows(lines: list[str]) -> list[tuple[str, str, str]]:
    lines = _repair_ocr_lines("\n".join(lines)).splitlines()
    header_rows = _extract_rows_by_price_sheet_header(lines)
    if header_rows:
        return header_rows

    rows: list[tuple[str, str, str]] = []
    for line in lines:
        if _is_noise_line(line):
            continue
        if any(keyword in line for keyword in SKIP_LINE_KEYWORDS):
            continue
        parsed = _parse_size_price_cells(_split_table_cells(line))
        if parsed:
            rows.append(parsed)
    return rows


def _extract_rows_by_price_sheet_header(lines: list[str]) -> list[tuple[str, str, str]]:
    header_index, column_map = _find_price_sheet_header(lines)
    if header_index is None:
        return []

    rows: list[tuple[str, str, str]] = []
    size_idx = column_map.get("size", 0)
    price_idx = column_map.get("price", 1)
    weight_idx = column_map.get("weight", 2)

    for line in lines[header_index + 1 :]:
        if _is_noise_line(line):
            continue
        if any(keyword in line for keyword in SKIP_LINE_KEYWORDS):
            continue
        cells = _split_table_cells(line)
        if len(cells) <= size_idx:
            continue
        size = _normalize_size_text(cells[size_idx].strip())
        if not _matches_size(size):
            continue
        price = cells[price_idx].strip() if price_idx < len(cells) else ""
        weight = cells[weight_idx].strip() if weight_idx < len(cells) else ""
        price, weight = _normalize_price_weight(price, weight)
        rows.append((size, price, weight))
    if rows and not any(price for _size, price, _weight in rows):
        return []
    return rows


def _find_price_sheet_header(lines: list[str]) -> tuple[int | None, dict[str, int]]:
    for index, line in enumerate(lines[:20]):
        cells = _split_table_cells(line)
        column_map: dict[str, int] = {}
        for cell_index, cell in enumerate(cells):
            field = _match_price_sheet_column(cell)
            if field and field not in column_map:
                column_map[field] = cell_index
        if "size" in column_map and "price" in column_map:
            if "weight" not in column_map:
                used = set(column_map.values())
                next_idx = max(used) + 1
                column_map["weight"] = next_idx
            return index, column_map
    return None, {}


def _match_price_sheet_column(cell: str) -> str | None:
    norm = re.sub(r"\s+", "", cell).lower()
    if not norm:
        return None
    if any(alias in norm for alias in ("产品尺寸",)) or norm == "尺寸":
        return "size"
    if any(alias in norm for alias in ("成本价",)) or norm in {"价格", "单价"}:
        return "price"
    if ("单品" in norm and "重量" in norm) or norm in {"重量", "重量kg", "重量(kg)"}:
        return "weight"
    return None


def _split_table_cells(line: str) -> list[str]:
    cleaned = line.rstrip()
    if not cleaned.strip():
        return []
    if "\t" in cleaned:
        return [part.strip() for part in cleaned.split("\t")]
    parts = re.split(r"\s{2,}", cleaned)
    if len(parts) >= 2:
        return [part.strip() for part in parts if part.strip()]
    compact = re.sub(r"\s+", " ", cleaned.strip())
    match = SIZE_PRICE_ROW.match(compact)
    if match:
        return [_normalize_size_text(match.group(1)), match.group(2), match.group(3) or ""]
    return [part.strip() for part in re.split(r"\s+", cleaned) if part.strip()]


def _parse_size_price_cells(cells: list[str]) -> tuple[str, str, str] | None:
    if not cells:
        return None
    size = ""
    numbers: list[str] = []
    for cell in cells:
        normalized_cell = _normalize_size_text(cell)
        if not size and _matches_size(normalized_cell):
            size = normalized_cell
            continue
        if re.fullmatch(r"\d+(?:\.\d+)?", cell):
            numbers.append(cell)
    if not size:
        match = SIZE_PRICE_ROW.match(" ".join(cells))
        if not match:
            return None
        price, weight = _normalize_price_weight(match.group(2), match.group(3) or "")
        return _normalize_size_text(match.group(1)), price, weight
    price = numbers[0] if len(numbers) >= 1 else ""
    weight = numbers[1] if len(numbers) >= 2 else ""
    price, weight = _normalize_price_weight(price, weight)
    return size, price, weight


def _normalize_price_weight(price: str, weight: str) -> tuple[str, str]:
    price_num = _parse_number(price)
    weight_num = _parse_number(weight)
    if price_num is None and weight_num is None:
        return price, weight
    if price_num is None:
        if weight_num is not None and weight_num < 3.5:
            return "", weight
        return weight, ""
    if weight_num is None:
        if price_num < 3.5:
            return "", price
        return price, ""
    if price_num < 3.5 and weight_num >= 3.5:
        return weight, price
    if price_num >= 3.5 and weight_num < 3.5:
        return price, weight
    if price_num < 5 and weight_num >= 5:
        return weight, price
    return price, weight


def _parse_number(value: str) -> float | None:
    text = str(value or "").strip()
    if not text:
        return None
    match = re.search(r"\d+(?:\.\d+)?", text)
    return float(match.group()) if match else None


def _rows_to_table_text(rows: list[list[str]]) -> str:
    header_index, header_map = _find_header_row(rows)
    if header_index is None or not header_map:
        return _plain_rows_to_text(rows)

    lines = ["\t".join(CANONICAL_HEADERS)]
    for row in rows[header_index + 1 :]:
        if _is_instruction_row(row):
            continue
        mapped = _map_row_to_canonical(row, header_map)
        if not any(mapped.values()):
            continue
        lines.append("\t".join(mapped.get(key, "") for key in CANONICAL_HEADERS))
    return "\n".join(lines) if len(lines) > 1 else ""


def _find_header_row(rows: list[list[str]]) -> tuple[int | None, dict[str, int]]:
    best_index: int | None = None
    best_map: dict[str, int] = {}
    best_score = 0
    for index, row in enumerate(rows[:20]):
        header_map = _header_map_for_row(row)
        score = len(header_map)
        if score > best_score:
            best_score = score
            best_index = index
            best_map = header_map
    if best_score >= 2:
        return best_index, best_map
    return None, {}


def _header_map_for_row(row: list[str]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index, cell in enumerate(row):
        normalized = _cell_text(cell).lower()
        if not normalized:
            continue
        for field, aliases in INPUT_HEADER_ALIASES.items():
            if field in header_map:
                continue
            if normalized in {alias.lower() for alias in aliases}:
                header_map[field] = index
                break
    return header_map


def _map_row_to_canonical(row: list[str], header_map: dict[str, int]) -> dict[str, str]:
    def value(field: str) -> str:
        index = header_map.get(field)
        if index is None or index >= len(row):
            return ""
        return _cell_text(row[index])

    return {
        "产品名称": value("product_name"),
        "颜色": value("color"),
        "尺寸": value("size"),
        "价格": value("price"),
        "重量": value("weight"),
        "供应商": value("supplier"),
    }


def _plain_rows_to_text(rows: list[list[str]]) -> str:
    lines: list[str] = []
    for row in rows:
        if _is_instruction_row(row):
            continue
        cleaned = [_cell_text(cell) for cell in row if _cell_text(cell)]
        if cleaned:
            lines.append("\t".join(cleaned))
    return "\n".join(lines)


def _is_instruction_row(row: list[str]) -> bool:
    joined = "".join(_cell_text(cell) for cell in row)
    return joined.startswith("导入说明") or "标红为必填" in joined


def _ocr_tokens(image: Any) -> list[tuple[str, float, tuple[float, float, float, float]]]:
    from app.ocr_utils import ocr_tokens

    return ocr_tokens(image)


def _token_center(bbox: tuple[float, float, float, float]) -> tuple[float, float]:
    return bbox[0] + bbox[2] / 2, bbox[1] + bbox[3] / 2


def _cluster_x_centers(values: list[float], gap: float = 0.10) -> list[float]:
    if not values:
        return []
    sorted_values = sorted(values)
    clusters: list[list[float]] = [[sorted_values[0]]]
    for value in sorted_values[1:]:
        if value - clusters[-1][-1] > gap:
            clusters.append([value])
        else:
            clusters[-1].append(value)
    return [sum(cluster) / len(cluster) for cluster in clusters]


def _average_numeric_values(texts: list[str]) -> float | None:
    numbers = [num for num in (_parse_number(text) for text in texts) if num is not None]
    if not numbers:
        return None
    return sum(numbers) / len(numbers)


def _infer_price_weight_column_centers(
    numeric_items: list[tuple[str, float, float]],
) -> tuple[float | None, float | None]:
    if not numeric_items:
        return None, None
    cluster_centers = _cluster_x_centers([x_center for _text, x_center, _y_center in numeric_items])
    if not cluster_centers:
        return None, None
    if len(cluster_centers) == 1:
        center = cluster_centers[0]
        return center, center

    left_center, right_center = cluster_centers[0], cluster_centers[-1]
    tolerance = 0.08
    left_texts = [text for text, x_center, _y in numeric_items if abs(x_center - left_center) <= tolerance]
    right_texts = [text for text, x_center, _y in numeric_items if abs(x_center - right_center) <= tolerance]
    left_avg = _average_numeric_values(left_texts)
    right_avg = _average_numeric_values(right_texts)
    if left_avg is not None and right_avg is not None:
        if left_avg < right_avg and left_avg < 5:
            return right_center, left_center
        if right_avg < left_avg and right_avg < 5:
            return left_center, right_center
    return left_center, right_center


def _assign_price_weight_column(x_center: float, price_center: float, weight_center: float) -> str:
    if abs(x_center - price_center) <= abs(x_center - weight_center):
        return "price"
    return "weight"


def _nearest_value_by_y(
    y_target: float,
    items: list[tuple[str, float]],
    used_indices: set[int],
    tolerance: float = 0.03,
) -> str:
    best_index: int | None = None
    best_distance = float("inf")
    for index, (value, y_center) in enumerate(items):
        if index in used_indices:
            continue
        distance = abs(y_center - y_target)
        if distance < best_distance:
            best_distance = distance
            best_index = index
    if best_index is None or best_distance > tolerance:
        return ""
    used_indices.add(best_index)
    return items[best_index][0]


def _column_major_meta_line(cleaned: str) -> bool:
    if re.fullmatch(r"\d+(?:\.\d+)?", cleaned):
        return False
    return any(
        keyword in cleaned
        for keyword in ("忻慕", "颜色", "供应商", "门幅", "度2米", "沙发垫", "定制", "平方")
    )


def _column_major_size_threshold(size_x_centers: list[float], numeric_x_centers: list[float]) -> float:
    if size_x_centers:
        return max(size_x_centers) + 0.06
    if numeric_x_centers:
        return min(numeric_x_centers) - 0.06
    return 0.25


def _repair_size_ocr_typo(text: str) -> str:
    cleaned = str(text or "").strip()
    if not cleaned:
        return ""
    cleaned = cleaned.replace("村", "*").replace("×", "*").replace("］", "套").replace("]", "套")
    cleaned = re.sub(r"(\d+)Q$", r"\g<1>0", cleaned, flags=re.I)
    if re.fullmatch(r"\d{6}", cleaned):
        cleaned = f"{cleaned[:2]}*{cleaned[2:]}"
    if re.fullmatch(r"\d{5}", cleaned) and "*" not in cleaned:
        cleaned = f"{cleaned[:2]}*{cleaned[2:]}"
    return _normalize_size_text(cleaned)


def _looks_like_column_major_layout(tokens: list[tuple[str, float, tuple[float, float, float, float]]]) -> bool:
    size_count = 0
    numeric_count = 0
    for text, _confidence, bbox in tokens:
        cleaned = str(text).strip()
        if not cleaned or _column_major_meta_line(cleaned):
            continue
        x_center, _y_center = _token_center(bbox)
        if _matches_size(_repair_size_ocr_typo(cleaned)):
            size_count += 1
        elif re.fullmatch(r"\d+(?:\.\d+)?", cleaned):
            numeric_count += 1
    return size_count >= 8 and numeric_count >= 6


def _tokens_to_column_major_table(tokens: list[tuple[str, float, tuple[float, float, float, float]]]) -> str:
    meta_lines: list[str] = []
    size_items: list[tuple[str, float, float]] = []
    numeric_items: list[tuple[str, float, float]] = []
    size_x_centers: list[float] = []

    for text, _confidence, bbox in tokens:
        cleaned = str(text).strip()
        if not cleaned or _is_noise_line(cleaned):
            continue
        x_center, y_center = _token_center(bbox)
        if _column_major_meta_line(cleaned):
            meta_lines.append(cleaned)
            continue
        size = _repair_size_ocr_typo(cleaned)
        if _matches_size(size):
            size_items.append((size, x_center, y_center))
            size_x_centers.append(x_center)
            continue
        if re.fullmatch(r"\d+(?:\.\d+)?", cleaned):
            numeric_items.append((cleaned, x_center, y_center))

    if not size_items:
        return ""

    size_threshold = _column_major_size_threshold(size_x_centers, [x for _t, x, _y in numeric_items])
    table_numeric_items = [
        item for item in numeric_items if item[1] > size_threshold + 0.02
    ]
    price_center, weight_center = _infer_price_weight_column_centers(table_numeric_items)
    if price_center is None or weight_center is None:
        return ""

    price_items: list[tuple[str, float]] = []
    weight_items: list[tuple[str, float]] = []
    for cleaned, x_center, y_center in table_numeric_items:
        column = _assign_price_weight_column(x_center, price_center, weight_center)
        if column == "price":
            price_items.append((cleaned, y_center))
        else:
            weight_items.append((cleaned, y_center))

    size_items.sort(key=lambda item: -item[2])
    price_items.sort(key=lambda item: -item[1])
    weight_items.sort(key=lambda item: -item[1])

    used_prices: set[int] = set()
    used_weights: set[int] = set()
    data_lines: list[str] = []
    for size, _size_x, size_y in size_items:
        price = _nearest_value_by_y(size_y, price_items, used_prices)
        weight = _nearest_value_by_y(size_y, weight_items, used_weights)
        if not price and not weight and price_items and weight_items:
            price = price_items[min(len(used_prices), len(price_items) - 1)][0]
            weight = weight_items[min(len(used_weights), len(weight_items) - 1)][0]
        if price or weight:
            price, weight = _normalize_price_weight(price, weight)
        data_lines.append("\t".join([size, price, weight]))

    lines = meta_lines + ["产品尺寸\t成本价\t单品重量"] + data_lines
    return "\n".join(lines)


def _column_major_has_weight_data(text: str) -> bool:
    for line in text.splitlines():
        cells = _split_table_cells(line)
        if len(cells) < 3 or not _matches_size(_normalize_size_text(cells[0])):
            continue
        if str(cells[2] or "").strip():
            return True
    return False


def _tokens_to_table_text(tokens: list[tuple[str, float, tuple[float, float, float, float]]]) -> str:
    merged = _repair_ocr_lines(_tokens_to_merged_row_text(tokens))
    if _looks_like_column_major_layout(tokens):
        column_major = _repair_ocr_lines(_tokens_to_column_major_table(tokens))
        if column_major.strip() and _column_major_has_weight_data(column_major):
            return column_major

    column_text = _repair_ocr_lines(_tokens_to_column_table(tokens))

    if column_text.strip() and _ocr_text_has_price_data(column_text):
        return column_text
    if merged.strip():
        return merged

    if len(tokens) == 1 or all(len(token) == 3 and isinstance(token[0], str) and token[2][2] == 1.0 for token in tokens):
        lines = [token[0] for token in tokens]
        return _repair_ocr_lines("\n".join(_normalize_ocr_line(line) for line in lines if line.strip()))

    return column_text or merged


def _tokens_to_merged_row_text(tokens: list[tuple[str, float, tuple[float, float, float, float]]]) -> str:
    rows: dict[int, list[tuple[float, str]]] = {}
    for text, _confidence, bbox in tokens:
        cleaned = str(text).strip()
        if not cleaned:
            continue
        y_center = bbox[1] + bbox[3] / 2
        y_key = int(round(y_center * 80))
        x_center = bbox[0] + bbox[2] / 2
        rows.setdefault(y_key, []).append((x_center, cleaned))

    lines: list[str] = []
    for _y_key in sorted(rows):
        cells = [cell for _x, cell in sorted(rows[_y_key], key=lambda item: item[0])]
        if cells:
            lines.append(_format_ocr_row_line(cells))
    return "\n".join(_normalize_ocr_line(line) for line in lines if line)


def _format_ocr_row_line(cells: list[str]) -> str:
    size_idx = next((index for index, cell in enumerate(cells) if _matches_size(cell)), None)
    if size_idx is None:
        return "\t".join(cells)

    size = _normalize_size_text(cells[size_idx])
    before = cells[:size_idx]
    after = cells[size_idx + 1 :]
    numbers = [cell for cell in after if re.fullmatch(r"\d+(?:\.\d+)?", cell)]
    non_numbers = [cell for cell in after if not re.fullmatch(r"\d+(?:\.\d+)?", cell)]
    price = numbers[0] if numbers else ""
    weight = numbers[1] if len(numbers) > 1 else ""
    if price or weight:
        price, weight = _normalize_price_weight(price, weight)
    row_parts = before + [size, price, weight] + non_numbers
    return "\t".join(row_parts)


def _tokens_to_column_table(tokens: list[tuple[str, float, tuple[float, float, float, float]]]) -> str:
    items = [(str(text).strip(), bbox) for text, _confidence, bbox in tokens if str(text).strip()]
    if len(items) < 6:
        return ""

    header_items = [(text, bbox) for text, bbox in items if _match_price_sheet_column(text)]
    if len(header_items) < 2:
        return ""

    columns = sorted(header_items, key=lambda item: item[1][0])
    column_centers = [bbox[0] + bbox[2] / 2 for _text, bbox in columns]
    column_count = len(column_centers)
    if column_count < 2:
        return ""

    def assign_column(x_center: float) -> int:
        return min(range(column_count), key=lambda idx: abs(column_centers[idx] - x_center))

    grid: dict[tuple[int, int], list[tuple[float, str]]] = {}
    for text, bbox in items:
        x_center = bbox[0] + bbox[2] / 2
        y_center = bbox[1] + bbox[3] / 2
        col = assign_column(x_center)
        row = int(round(y_center * 200))
        grid.setdefault((row, col), []).append((x_center, text))

    lines: list[str] = []
    for row in sorted({key[0] for key in grid}):
        cells: list[str] = []
        for col in range(column_count):
            parts = [text for _x, text in sorted(grid.get((row, col), []), key=lambda item: item[0])]
            cells.append(" ".join(parts) if parts else "")
        if any(cell.strip() for cell in cells):
            lines.append("\t".join(cells))
    return "\n".join(lines)


def _normalize_ocr_line(line: str) -> str:
    if "\t" in line:
        return line
    if "," in line and line.count(",") >= 2:
        return "\t".join(part.strip() for part in line.split(","))
    compact = re.sub(r"\s+", " ", line.strip())
    size_price_match = SIZE_PRICE_ROW.match(compact)
    if size_price_match:
        price, weight = _normalize_price_weight(size_price_match.group(2), size_price_match.group(3) or "")
        return "\t".join([size_price_match.group(1), price, weight])
    parts = re.split(r"\s{2,}", line.strip())
    if len(parts) >= 3:
        return "\t".join(parts)
    space_parts = re.split(r"\s+", line.strip())
    if (
        len(space_parts) == 3
        and _matches_size(space_parts[0])
        and re.fullmatch(r"\d+(?:\.\d+)?", space_parts[1])
        and re.fullmatch(r"\d+(?:\.\d+)?", space_parts[2])
    ):
        price, weight = _normalize_price_weight(space_parts[1], space_parts[2])
        return "\t".join([_normalize_size_text(space_parts[0]), price, weight])
    if _looks_like_standard_row(line):
        return "\t".join(part.strip() for part in re.split(r"\s+", line.strip()))
    return line.strip()


def _looks_like_standard_row(line: str) -> bool:
    parts = re.split(r"\s+", line.strip())
    if len(parts) < 4:
        return False
    if parts[0] in {"产品名称", "产品名", "颜色", "尺寸"}:
        return True
    return bool(re.search(r"\d+\*\d+", parts[2] if len(parts) > 2 else ""))


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _suffix(filename: str) -> str:
    match = re.search(r"(\.[^.]+)$", filename.lower())
    return match.group(1) if match else ""
