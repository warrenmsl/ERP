from __future__ import annotations

from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import Any

DISTRIBUTION_IMPORT_INSTRUCTION = (
    "导入说明:\n"
    "1、标红为必填项：款式编码、商品编码、商品名\n"
    "2、基本售价、重量、库存、市场|吊牌价：只能输入数字\n"
    "3、若聚水潭ERP存在相同商品编码的商品，“基本售价”以ERP为准"
)

DISTRIBUTION_HEADERS = (
    "款式编码",
    "商品编码",
    "颜色",
    "规格",
    "商品主图",
    "商品详情图",
    "图片地址",
    "商品名称",
    "推荐文案",
    "商品描述",
    "宝贝链接",
    "库存",
    "重量(kg)",
    "基本售价",
    "市场|吊牌价",
    "最低分销控价",
    "最高分销控价",
    "供应商名",
)

NEW_PRODUCT_HEADERS = (
    "商品编码",
    "重量",
    "基本售价",
    "成本价",
    "其它属性1",
)


def build_distribution_xlsx(built: dict[str, Any]) -> bytes:
    rows = [_distribution_data_row(row) for row in built.get("rows", [])]
    return _workbook_bytes(
        DISTRIBUTION_HEADERS,
        rows,
        "商品导入模板",
        preamble_row=(DISTRIBUTION_IMPORT_INSTRUCTION,),
    )


def _distribution_data_row(row: dict[str, Any]) -> list[Any]:
    return [
        row.get("style_code", ""),
        row.get("sku_code", ""),
        row.get("erp_color", ""),
        row.get("erp_spec", ""),
        None,
        None,
        None,
        row.get("erp_goods_name", ""),
        None,
        None,
        None,
        None,
        row.get("weight") if row.get("weight_present") else None,
        row.get("base_selling_price") if row.get("price_present") else None,
        None,
        None,
        None,
        row.get("supplier", ""),
    ]


def build_new_product_xlsx(built: dict[str, Any]) -> bytes:
    rows = [
        [
            row.get("sku_code", ""),
            row.get("weight") if row.get("weight_present") else "",
            row.get("base_selling_price") if row.get("price_present") else "",
            _export_cost_price_one_decimal(row.get("cost_price")) if row.get("price_present") else "",
            row.get("supplier", ""),
        ]
        for row in built.get("rows", [])
    ]
    return _workbook_bytes(
        NEW_PRODUCT_HEADERS,
        rows,
        "Sheet1",
        column_formats={4: "0.0"},
    )


def _export_cost_price_one_decimal(value: Any) -> float:
    quantized = Decimal(str(float(value))).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
    return float(quantized)


def _workbook_bytes(
    headers: tuple[str, ...],
    rows: list[list[Any]],
    sheet_name: str,
    *,
    preamble_row: tuple[Any, ...] | None = None,
    column_formats: dict[int, str] | None = None,
) -> bytes:
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    data_start_row = 1
    if preamble_row is not None:
        preamble_cells = list(preamble_row) + [None] * max(0, len(headers) - len(preamble_row))
        sheet.append(preamble_cells[: len(headers)])
        data_start_row = 2
    sheet.append(list(headers))
    data_start_row += 1
    for row in rows:
        cells = list(row) + [None] * max(0, len(headers) - len(row))
        sheet.append(cells[: len(headers)])
    if column_formats:
        for column_index, number_format in column_formats.items():
            for row_index in range(data_start_row, sheet.max_row + 1):
                sheet.cell(row=row_index, column=column_index).number_format = number_format
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
